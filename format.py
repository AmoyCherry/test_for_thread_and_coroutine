from fileinput import filename
from openpyxl import Workbook, load_workbook

all_data = []
# 读取的原始数据文件名
file_name = 'user_cor_4096B.txt'
raw_file = open(file_name)
line = raw_file.readline()
while line:
    if line[0] == '#':
        ret = 0
        while line != '\n':
            ret = line
            line = raw_file.readline()
        #print(ret)
        all_data.append(int(ret.strip('\n')))
    else:
        line = raw_file.readline()
    
raw_file.close()


# creat a workbook obj
wb = Workbook()

sheet = wb.active

index = 0

path = 'empty.xlsx'
tp_wb = load_workbook(path)
tp_sheet = (tp_wb.worksheets)[0]
for i in range(2, 25):
    for j in range(2, 22):
        (length, num) = (tp_sheet.cell(i, 1).value, tp_sheet.cell(1, j).value)
        if num >= length and num % length == 0:
            print([length, num])
            tp_sheet.cell(i, j).value = all_data[index]
            index += 1

result_name = file_name[:-3] + 'xlsx'
tp_wb.save(result_name)




